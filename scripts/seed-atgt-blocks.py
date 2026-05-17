#!/usr/bin/env python3
"""
TrishTEAM Phase 42 wave 9.1 — Seed Firestore /atgt_blocks từ database.xlsx.

Đọc file Excel "database.xlsx" (415 tài sản ATGT) → upload Firestore collection /atgt_blocks.

Usage:
    cd scripts
    pip install openpyxl firebase-admin --user
    python seed-atgt-blocks.py path/to/database.xlsx
    python seed-atgt-blocks.py path/to/database.xlsx --dry-run   # chỉ in ra, không upload

Yêu cầu:
    - File service-account.json đặt tại scripts/service-account.json (gốc từ Firebase Console)
    - Firestore project: trishteam-17c2d

Field mapping (Excel column → Firestore field):
    Tên file       → fileName, id (slug từ tên)
    Dạng địa vật   → shapeKind ('Block' → 'block', 'Linetype' → 'linetype')
    Tên địa vật    → label
    Ý nghĩa        → meaning
    Hướng          → orientation ('Vuông góc' → 'perpendicular', 'Song song' → 'parallel')
    Loại tài sản   → category
    Ghi chú        → description
"""
import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print('ERROR: pip install openpyxl --user', file=sys.stderr)
    sys.exit(1)


# Color map theo "Loại tài sản" (ACI AutoCAD)
CATEGORY_COLOR = {
    'Biển báo':         1,    # Red
    'Vạch sơn':         2,    # Yellow
    'Đèn tín hiệu':     3,    # Green
    'Hộ lan mềm':       4,    # Cyan
    'Cọc tiêu':         5,    # Blue
    'Rãnh dọc':         6,    # Magenta
    'Cống ngang':       30,   # Orange
    'Tiêu phản quang':  7,    # White
    'Gương cầu lồi':    8,    # Gray
    'Lí trình':         9,    # Light gray
}


def slug(text: str) -> str:
    """Chuyển 'Tên file' (vd '12.BB' hoặc '0.LT') thành Firestore docId an toàn."""
    if not text:
        return ''
    # Giữ chữ + số + dấu chấm + dấu gạch, đổi khoảng trắng thành _
    s = str(text).strip().lower()
    s = re.sub(r'\s+', '_', s)
    s = re.sub(r'[^a-z0-9._-]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s


def map_shape_kind(raw: str) -> str:
    return 'linetype' if 'linetype' in str(raw or '').lower() else 'block'


def map_orientation(raw: str) -> str:
    return 'parallel' if 'song' in str(raw or '').lower() else 'perpendicular'


def parse_excel(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb['Database']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    print(f'Header: {header}')

    out = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        stt, ten_file, dang, ten_dia_vat, y_nghia, huong, loai, ghi_chu = (r + (None,) * 8)[:8]
        if not ten_file or not ten_dia_vat:
            continue
        doc_id = slug(str(ten_file))
        if not doc_id:
            continue
        category = str(loai or 'Khác').strip()
        block = {
            'id': doc_id,
            'fileName': str(ten_file).strip() + '.dwg' if not str(ten_file).lower().endswith('.dwg') else str(ten_file),
            'label': str(ten_dia_vat).strip(),
            'meaning': str(y_nghia or '').strip(),
            'shapeKind': map_shape_kind(dang),
            'orientation': map_orientation(huong),
            'category': category,
            'description': str(ghi_chu or '').strip(),
            'colorIndex': CATEGORY_COLOR.get(category, 7),
            'defaultScale': 1.0,
            'hatchName': '',
        }
        out.append(block)
    return out


def upload_firestore(blocks: list[dict], cred_path: Path) -> None:
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        print('ERROR: pip install firebase-admin --user', file=sys.stderr)
        sys.exit(1)

    if not cred_path.exists():
        print(f'ERROR: missing service account: {cred_path}', file=sys.stderr)
        print(' → download Firebase Admin SDK key từ console.firebase.google.com', file=sys.stderr)
        sys.exit(1)

    cred = credentials.Certificate(str(cred_path))
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred)
    db = firestore.client()
    coll = db.collection('atgt_blocks')

    print(f'\nUploading {len(blocks)} blocks to Firestore /atgt_blocks ...')
    import time
    now_ms = int(time.time() * 1000)
    written = 0
    batch = db.batch()
    batch_count = 0
    for b in blocks:
        b_with_ts = {**b, 'updated_at': now_ms}
        ref = coll.document(b['id'])
        batch.set(ref, b_with_ts, merge=True)
        batch_count += 1
        if batch_count >= 400:
            batch.commit()
            written += batch_count
            print(f'  Committed batch ({written} so far)')
            batch = db.batch()
            batch_count = 0
    if batch_count > 0:
        batch.commit()
        written += batch_count
    print(f'✅ Done. Wrote {written} docs.')


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', type=Path, help='Path tới database.xlsx')
    ap.add_argument('--dry-run', action='store_true', help='Chỉ in JSON ra console, KHÔNG upload Firestore')
    ap.add_argument('--cred', type=Path, default=Path(__file__).parent / 'service-account.json',
                    help='Path tới Firebase service account JSON (default scripts/service-account.json)')
    ap.add_argument('--limit', type=int, default=0, help='Chỉ xử lý N record đầu (debug)')
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f'ERROR: file không tồn tại: {args.xlsx}', file=sys.stderr)
        sys.exit(1)

    blocks = parse_excel(args.xlsx)
    print(f'Parsed {len(blocks)} blocks.')

    if args.limit > 0:
        blocks = blocks[:args.limit]
        print(f'Limited to first {len(blocks)} blocks.')

    # Group stats
    from collections import Counter
    print('\nCategory breakdown:')
    for cat, count in Counter(b['category'] for b in blocks).most_common():
        print(f'  {cat}: {count}')
    print('Shape kind:')
    for k, c in Counter(b['shapeKind'] for b in blocks).most_common():
        print(f'  {k}: {c}')
    print('Orientation:')
    for k, c in Counter(b['orientation'] for b in blocks).most_common():
        print(f'  {k}: {c}')

    if args.dry_run:
        print('\n--- DRY RUN — first 5 records ---')
        for b in blocks[:5]:
            print(json.dumps(b, ensure_ascii=False, indent=2))
        print(f'\n(dry-run) Would write {len(blocks)} docs. Re-run without --dry-run to upload.')
        return

    upload_firestore(blocks, args.cred)


if __name__ == '__main__':
    main()
